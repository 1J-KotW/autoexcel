import openpyxl
import json
import sys
from collections import defaultdict

# Column names (Russian)
MATERIAL_COL = "Материал"
UNIT_COL = "Единица измерения"
PRICE_COL = "Цена материала, за единицу"
LABOR_COL = "Стоимость работ, за единицу"
STATUS_COL = "Статус заполнения"

def load_catalog():
    try:
        with open('materials_catalog.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print("Catalog file not found. Please create materials_catalog.json")
        sys.exit(1)

def find_column_index(headers, col_name):
    try:
        return headers.index(col_name) + 1  # 1-based
    except ValueError:
        return None

def add_column_if_missing(ws, col_name, headers):
    if col_name not in headers:
        max_col = ws.max_column
        ws.cell(1, max_col + 1).value = col_name
        headers.append(col_name)
        return max_col + 1
    else:
        return find_column_index(headers, col_name)

def collect_local_catalog(wb):
    local_catalog = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        material_idx = find_column_index(headers, MATERIAL_COL)
        unit_idx = find_column_index(headers, UNIT_COL)
        price_idx = find_column_index(headers, PRICE_COL)
        labor_idx = find_column_index(headers, LABOR_COL)

        if not material_idx or not unit_idx or not price_idx or not labor_idx:
            continue

        for row in range(2, ws.max_row + 1):
            material = ws.cell(row, material_idx).value
            unit = ws.cell(row, unit_idx).value
            price = ws.cell(row, price_idx).value
            labor = ws.cell(row, labor_idx).value

            if material and unit and price is not None and labor is not None:
                key = (material, unit)
                if key not in local_catalog:
                    local_catalog[key] = (price, labor)
    return local_catalog

def process_excel(excel_file, catalog):
    wb = openpyxl.load_workbook(excel_file)
    local_catalog = collect_local_catalog(wb)
    missing = defaultdict(list)
    ambiguous = defaultdict(list)
    filled_from_local = 0
    filled_from_catalog = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        material_idx = find_column_index(headers, MATERIAL_COL)
        unit_idx = find_column_index(headers, UNIT_COL)

        if not material_idx or not unit_idx:
            print(f"Sheet '{sheet_name}' missing required columns '{MATERIAL_COL}' or '{UNIT_COL}'. Skipping.")
            continue

        price_idx = add_column_if_missing(ws, PRICE_COL, headers)
        labor_idx = add_column_if_missing(ws, LABOR_COL, headers)
        status_idx = add_column_if_missing(ws, STATUS_COL, headers)

        for row in range(2, ws.max_row + 1):
            material = ws.cell(row, material_idx).value
            unit = ws.cell(row, unit_idx).value

            if material and unit:
                key = (material, unit)
                source = None
                if key in local_catalog:
                    price_val, labor_val = local_catalog[key]
                    source = "локальный"
                    filled_from_local += 1
                else:
                    matches = [m for m in catalog if m['name'] == material and m['unit'] == unit]
                    if len(matches) == 1:
                        m = matches[0]
                        price_val, labor_val = m['price'], m['labor_cost']
                        source = "справочник"
                        filled_from_catalog += 1
                    elif len(matches) > 1:
                        ws.cell(row, status_idx).value = "Неоднозначно"
                        ambiguous[sheet_name].append((material, unit))
                        continue
                    else:
                        ws.cell(row, status_idx).value = "Не найдено"
                        missing[sheet_name].append((material, unit))
                        continue

                if not ws.cell(row, price_idx).value:
                    ws.cell(row, price_idx).value = price_val
                if not ws.cell(row, labor_idx).value:
                    ws.cell(row, labor_idx).value = labor_val
                ws.cell(row, status_idx).value = f"Заполнено ({source})"
            else:
                ws.cell(row, status_idx).value = "Нет данных"

    # Save updated file
    output_file = excel_file.replace('.xlsx', '_filled.xlsx')
    wb.save(output_file)
    print(f"Updated file saved as: {output_file}")
    print(f"Заполнено из локальных данных: {filled_from_local}")
    print(f"Заполнено из справочника: {filled_from_catalog}")

    # Print reports
    if missing:
        print("\nМатериалы, не найденные ни в файле, ни в справочнике:")
        for sheet, mats in missing.items():
            print(f"Лист '{sheet}': {len(mats)} позиций")
            # Show first 5
            for mat, unit in mats[:5]:
                print(f"  - {mat} ({unit})")
            if len(mats) > 5:
                print(f"  ... и ещё {len(mats) - 5}")

    if ambiguous:
        print("\nМатериалы с неоднозначным совпадением:")
        for sheet, mats in ambiguous.items():
            print(f"Лист '{sheet}':")
            for mat, unit in mats:
                print(f"  - {mat} ({unit})")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python fill_prices.py <excel_file.xlsx>")
        sys.exit(1)

    excel_file = sys.argv[1]
    catalog = load_catalog()
    process_excel(excel_file, catalog)
